"""
implement_FS_BBG_formulas_utils.py
"""

import warnings

import openpyxl

warnings.simplefilter("ignore")


# ================================================================
def create_excel_safely(output_path, data):
    """安全にExcelファイルを作成する関数"""

    # データの確認
    print(f"データ型: {type(data)}")
    print(f"データ内容: {data!r}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        # 新しいワークブックを作成
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = "Price"  # type: ignore

        # データを安全に書き込み
        if data is not None:
            if isinstance(data, (list, tuple)):
                print(f"リスト形式のデータを書き込み中 (要素数: {len(data)})")
                for i, value in enumerate(data):
                    try:
                        sheet[f"A{i+2}"] = value  # type: ignore
                    except Exception as e:
                        print(f"行{i+2}でエラー: {e}")
                        sheet[f"A{i+2}"] = str(value)  # type: ignore
            else:
                print("単一値を書き込み中")
                sheet["A2"] = data  # type: ignore
        else:
            print("データがNullです")
            sheet["A2"] = ""  # type: ignore

        # ワークブックの状態確認
        print(f"シート数: {len(book.worksheets)}")
        print(f"シート名: {[ws.title for ws in book.worksheets]}")

        # 保存
        book.save(output_path)
        print(f"ファイルが正常に作成されました: {output_path}")

        # ファイルサイズ確認
        print(f"ファイルサイズ: {output_path.stat().st_size} bytes")

    except Exception as e:
        print(f"Excel作成中にエラーが発生: {e}")
        raise


# ================================================================
def create_identifier_list(series, suffix):
    """識別子リストを効率的に作成"""
    return series.dropna().drop_duplicates().apply(lambda x: f'"{x} {suffix}"').tolist()


# ================================================================
def split_list_into_chunks(lst, chunk_size=500):
    """リストを指定サイズで分割"""
    for i in range(0, len(lst), chunk_size):
        yield lst[i : i + chunk_size]


# ================================================================
def create_bql_formula(identifier_list):
    """BQL数式を作成"""
    return f"=BQL.LIST({','.join(identifier_list)})"


# ================================================================
def create_excel_with_chunked_data(output_path, identifier_dict, chunk_size=500):
    """500銘柄ごとに分割してExcelファイルを作成"""

    output_path.parent.mkdir(parents=True, exist_ok=True)

    book = openpyxl.Workbook()
    # デフォルトシートを削除
    book.remove(book.active)  # type: ignore

    total_sheets = 0

    for identifier_type, identifier_list in identifier_dict.items():
        if not identifier_list:
            continue

        # リストを500個ずつに分割
        chunks = list(split_list_into_chunks(identifier_list, chunk_size))
        num_chunks = len(chunks)

        print(
            f"{identifier_type}: {len(identifier_list)}銘柄 -> {num_chunks}シートに分割"
        )

        for i, chunk in enumerate(chunks, 1):
            # シート名を作成
            if num_chunks == 1:
                sheet_name = identifier_type
            else:
                sheet_name = f"{identifier_type}_{i}"

            # シートを作成
            sheet = book.create_sheet(sheet_name)

            # BQL数式を作成
            bql_formula = create_bql_formula(chunk)

            # ヘッダー情報
            sheet["A1"] = f"{identifier_type} BQL Formula"
            sheet["B1"] = f"Count: {len(chunk)}"
            sheet["C1"] = f"Sheet {i}/{num_chunks}"

            # BQL数式を配置
            sheet["A2"] = bql_formula

            # デバッグ情報
            sheet["A3"] = f"Formula length: {len(bql_formula)} characters"

            total_sheets += 1

            print(
                f"  シート '{sheet_name}': {len(chunk)}銘柄, 数式長: {len(bql_formula)}文字"
            )

    # サマリーシートを追加
    summary_sheet = book.create_sheet("Summary", 0)  # 最初に配置
    summary_sheet["A1"] = "BQL Formula Summary"
    summary_sheet["A2"] = f"Total sheets: {total_sheets}"
    summary_sheet["A3"] = f"Max items per sheet: {chunk_size}"
    summary_sheet["A4"] = "Created sheets:"

    row = 5
    for sheet_name in book.sheetnames[1:]:  # Summaryシート以外
        summary_sheet[f"A{row}"] = sheet_name
        row += 1

    book.save(output_path)
    print(f"\nExcelファイル作成完了: {output_path}")
    print(f"総シート数: {len(book.sheetnames)} (サマリーシート含む)")


# ================================================================


# ================================================================
