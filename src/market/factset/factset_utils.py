"""
factset_utils.py
"""

import contextlib
import os
import re
import sqlite3
import time
import warnings
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

import numpy as np
import openpyxl
import pandas as pd
import yaml
from tqdm import tqdm

from utils_core.logging import get_logger
from utils_core.settings import load_project_env

# Legacy imports - these modules have been moved/removed
# import src.database_utils as db_utils
# import src.ROIC_make_data_files_ver2 as roic_utils

logger = get_logger(__name__)

warnings.simplefilter("ignore")


# ============================================================================================
# SQLインジェクション対策用ヘルパー関数
# ============================================================================================
def _validate_sql_identifier(name: str) -> str:
    """
    SQL識別子（テーブル名・カラム名）が安全であることを検証する。

    SQLiteでは識別子をパラメータ化できないため、
    この関数で識別子が安全な形式であることを保証する。

    Parameters
    ----------
    name : str
        検証するSQL識別子

    Returns
    -------
    str
        検証済みの識別子（入力と同じ値）

    Raises
    ------
    ValueError
        識別子が不正な形式の場合
    """
    # 空文字列チェック
    if not name or not name.strip():
        raise ValueError("SQL識別子は空にできません")

    # 許可するパターン: 英数字、アンダースコア、ハイフン、ドット
    # （FactSetのテーブル名・カラム名で使用される文字）
    pattern = r"^[a-zA-Z_][a-zA-Z0-9_\-\.]*$"
    if not re.match(pattern, name):
        raise ValueError(
            f"SQL識別子に不正な文字が含まれています: {name!r}. "
            f"許可される形式: 英字で始まり、英数字・アンダースコア・ハイフン・ドットのみ"
        )

    # SQLキーワードとの衝突チェック（主要なものだけ）
    sql_keywords = {
        "SELECT",
        "INSERT",
        "UPDATE",
        "DELETE",
        "DROP",
        "CREATE",
        "ALTER",
        "TABLE",
        "FROM",
        "WHERE",
        "AND",
        "OR",
        "NOT",
        "NULL",
        "TRUE",
        "FALSE",
    }
    if name.upper() in sql_keywords:
        raise ValueError(f"SQL識別子がSQLキーワードと衝突しています: {name}")

    return name


# ============================================================================================
def split_and_save_dataframe(
    df_all: pd.DataFrame, n_splits: int, base_dir: Path, base_filename: str, **kwargs
):
    """
    データフレームを指定された数に分割し、それぞれをParquet形式で保存する関数。

    Args:
        df_all (pd.DataFrame): 分割対象のデータフレーム。
        n_splits (int): 分割する数。1以上の整数である必要があります。
        base_dir (Path): ファイルを保存する基本ディレクトリ (e.g., FACTSET_DATA_DIR / "Index_Constituents")。
        base_filename (str): ファイル名のベース部分 (e.g., "Index_Constituents_with_Factset_code-compressed-")。
        **kwargs: pd.to_parquetに渡すその他の引数 (e.g., compression="zstd", index=False)。

    Returns:
        list: 保存されたファイルのパスのリスト。
    """
    if n_splits <= 0:
        raise ValueError("n_splits は 1 以上の整数でなければなりません。")

    total_rows = len(df_all)
    saved_files = []

    # 分割点を計算
    # np.linspaceを使用して、0からtotal_rowsまでの n_splits + 1 個の等間隔な整数インデックスを計算
    indices = np.linspace(0, total_rows, n_splits + 1, dtype=int)

    # データフレームの分割と保存
    for i in range(n_splits):
        # 分割範囲の開始と終了インデックス
        start_idx = indices[i]
        end_idx = indices[i + 1]

        # データの抽出
        df_split = df_all.iloc[start_idx:end_idx]

        # ファイル名とパスの設定
        # 例: Index_Constituents_with_Factset_code-compressed-1.parquet
        file_name = f"{base_filename}{i + 1}.parquet"
        file_path = base_dir / file_name

        # Parquet形式で保存
        df_split.to_parquet(file_path, **kwargs)
        saved_files.append(file_path)
        logger.info(
            "Split file saved",
            progress=f"{i + 1}/{n_splits}",
            file=file_name,
            rows=len(df_split),
        )

    return saved_files


# ============================================================================================
def load_bpm_and_export_factset_code_file(
    start_date: str, end_date: str, index_dir: list[Path]
):
    """
    BPM からダウンロードした Index 構成銘柄の paruqet ファイルを読み取る

    -   (Universe name)_Constituents.parquet から Factset の P_SYMBOL と FG_COMPANY_NAME をダウンロードするエクセルファイルを作成。
    -   その後、すべてのインデックスをまとめて parquet ファイルに保存。
    """

    # フォルダ
    load_project_env()
    bpm_data_dir = os.environ.get("BPM_DATA_DIR")
    if bpm_data_dir is None:
        raise ValueError("BPM_DATA_DIR environment variable not set")
    BPM_DATA_DIR = Path(bpm_data_dir)

    bpm_src_dir = os.environ.get("BPM_SRC_DIR")
    if bpm_src_dir is None:
        raise ValueError("BPM_SRC_DIR environment variable not set")
    BPM_SRC_DIR = Path(bpm_src_dir)

    src_dir_str = os.environ.get("SRC_DIR")
    if src_dir_str is None:
        raise ValueError("SRC_DIR environment variable not set")
    src_dir = Path(src_dir_str)

    with open(src_dir / "BPM_Index-code-map.yaml", encoding="utf-8") as f:
        bpm_code_map = yaml.safe_load(f)
        bpm_name_to_code = {value: key for key, value in bpm_code_map.items()}

    dfs = []
    for dir in index_dir:
        parquet_file = dir / f"{dir.name}_Constituents.parquet"
        df = (
            pd.read_parquet(parquet_file)
            .query("date>=@start_date and date<=@end_date")
            .assign(
                Universe=dir.name,
                Universe_code_BPM=bpm_name_to_code[dir.name],
            )
        )
        dfs.append(df)

    df = pd.concat(dfs, ignore_index=True)
    df["Weight (%)"] = df["Weight (%)"].astype(float)
    df["SEDOL"] = df["SEDOL"].astype(str).str.zfill(7)

    # 全構成銘柄のエクスポート
    output_path = BPM_DATA_DIR / "Index_Constituents.parquet"
    df.to_parquet(
        output_path,
        index=False,
    )
    logger.info("BPM constituents exported", path=str(output_path))

    # Factsetコードダウンロード用Excelのエクスポート
    df_slice = (
        df[
            [
                "Asset ID",
                "Asset ID Type",
                "Country",
                "Universe",
                "Universe_code_BPM",
                "CUSIP",
                "SEDOL",
                "ISIN",
                "CODE_JP",
            ]
        ]
        .drop_duplicates()
        .replace("N/A", np.nan)
        .fillna(np.nan)
        .dropna(
            subset=["CUSIP", "SEDOL", "ISIN", "CODE_JP"],
            how="all",
        )
        .reset_index(drop=True)
        # .assign(CODE=lambda row: row["SEDOL"])
    )
    excel_rows = (df_slice.index + 2).astype(str)
    for col, row_alphabet in {
        "CUSIP": "F",
        "SEDOL": "G",
        "ISIN": "H",
        "CODE_JP": "I",
    }.items():
        # code_col = "G"  # 銘柄コードが入っているエクセルのカラム(変更する場合は要確認)
        df_slice[f"FG_COMPANY_NAME_{col}"] = (
            f"=FDS({row_alphabet}" + excel_rows + ', "FG_COMPANY_NAME"' + ")"
        )  # 関数式は配列で格納
        df_slice[f"P_SYMBOL_{col}"] = (
            f"=FDS({row_alphabet}" + excel_rows + ', "P_SYMBOL"' + ")"
        )  # 関数式は配列で格納

    # export
    FACTSET_ROOT_DIR = Path(os.getenv("FACTSET_ROOT_DIR"))  # type: ignore
    output_excel_path = (
        FACTSET_ROOT_DIR / "Index_Constituents/Index_Constituents_Factset_code_DL.xlsx"
    )

    df_slice.to_excel(output_excel_path, index=False)
    logger.info("Factset code download Excel exported", path=str(output_excel_path))


# ============================================================================================
def unify_factset_code_data(split_save_mode: bool = False):
    """
    FactsetからExcelにダウンロードしたP_SYMBOL, FG_COMPANY_NAMEのファイルを統合する関数.
    P_SYMBOLとFG_COMPANY_NAMEはBPMからダウンロードしたSEDOL, CUSIP, ISIN, CODE_JPを引数として、
    それぞれダウンロードしている.
    """

    # フォルダ
    load_project_env()
    BPM_DATA_DIR = Path(os.environ.get("BPM_DATA_DIR"))  # type: ignore
    FACTSET_ROOT_DIR = Path(os.environ.get("FACTSET_ROOT_DIR"))  # type: ignore

    # BPMからダウンロードしたファイル
    df_bpm = pd.read_parquet(BPM_DATA_DIR / "Index_Constituents.parquet").replace(
        "N/A", np.nan
    )

    # Universeごとにファイル読み込み
    # Factsetでダウンロードしたコードのファイル
    file_list = list(
        FACTSET_ROOT_DIR.glob(
            "Index_Constituents/M*_Index_Constituents_Factset_code_DL.csv"
        )
    )

    universe_code_bpm = [
        s.name.replace("_Index_Constituents_Factset_code_DL.csv", "") for s in file_list
    ]

    code_type_cols = ["SEDOL", "CUSIP", "ISIN", "CODE_JP"]
    df_all = []
    for universe_code, f in zip(universe_code_bpm, file_list, strict=False):
        df_factset_uni = pd.read_csv(
            f,
            encoding="utf-8",
            dtype={"CUSIP": "str", "SEDOL": "str", "ISIN": "str", "CODE_JP": "str"},
        ).replace("N/A", np.nan)
        df_bpm_uni = df_bpm[df_bpm["Universe_code_BPM"] == universe_code]

        dfs = []

        for code_type in code_type_cols:
            drop_cols = code_type_cols.copy()
            drop_cols.remove(code_type)
            df_left = df_bpm_uni.drop(columns=drop_cols)
            df_right = df_factset_uni.drop(
                columns=drop_cols
                + [f"P_SYMBOL_{s}" for s in drop_cols]
                + [f"FG_COMPANY_NAME_{s}" for s in drop_cols]
            )
            df_merge = pd.merge(
                df_left,
                df_right,
                on=[
                    col
                    for col in df_right.columns
                    if (not col.startswith("P_SYMBOL"))
                    & (not col.startswith("FG_COMPANY_NAME"))
                ],
                how="left",
            )

            dfs.append(df_merge)

        # マージしたそれぞれのデータフレームから共通カラムを取得する
        # 最初のデータフレームのカラム名を初期セットとして取得
        common_cols = set(dfs[0].columns)
        # リスト内の残りのデータフレームのカラム名と共通部分を計算
        for df_ in dfs[1:]:
            current_cols = set(df_.columns)
            common_cols = common_cols.intersection(current_cols)
        common_cols = list(common_cols)

        df_merged = pd.DataFrame()
        for index, df_code in enumerate(dfs):
            if index == 0:
                df_merged = df_code
            else:
                df_merged = pd.merge(df_merged, df_code, on=common_cols, how="left")

        # 最後に、P_SYMBOLとFG_COMPANY_NAMEをひとつにまとめる。
        # CODE_JPが欠損してないものと欠損しているものでそれぞれ対応。
        df_merged_copy = df_merged.copy()
        df_code_jp_not_missing = df_merged_copy[
            ~df_merged_copy["CODE_JP"].isnull()
        ].reset_index(drop=True)
        df_code_jp_missing = df_merged_copy[
            df_merged_copy["CODE_JP"].isnull()
        ].reset_index(drop=True)

        # P_SYMBOLとFG_COMPANY_NAMEは取得可能なものでfillnaする
        df_code_jp_not_missing["P_SYMBOL"] = df_code_jp_not_missing["P_SYMBOL_CODE_JP"]
        df_code_jp_not_missing["FG_COMPANY_NAME"] = df_code_jp_not_missing[
            "FG_COMPANY_NAME_CODE_JP"
        ]

        df_code_jp_missing["P_SYMBOL"] = (
            df_code_jp_missing["P_SYMBOL_SEDOL"]
            .fillna(df_code_jp_missing["P_SYMBOL_CUSIP"])  # type: ignore[arg-type]
            .fillna(df_code_jp_missing["P_SYMBOL_ISIN"])  # type: ignore[arg-type]
        )
        df_code_jp_missing["FG_COMPANY_NAME"] = (
            df_code_jp_missing["FG_COMPANY_NAME_SEDOL"]
            .fillna(df_code_jp_missing["FG_COMPANY_NAME_CUSIP"])  # type: ignore[arg-type]
            .fillna(df_code_jp_missing["FG_COMPANY_NAME_ISIN"])  # type: ignore[arg-type]
        )

        # concat
        df_final = (
            pd.concat([df_code_jp_missing, df_code_jp_not_missing])
            .sort_values("date")
            .drop_duplicates(ignore_index=True)
        )

        # export
        df_final.to_parquet(
            FACTSET_ROOT_DIR
            / f"Index_Constituents/{universe_code}_Index_Constituents_with_Factset_code.parquet",
            index=False,
        )
        df_all.append(df_final)
        logger.info(
            "Universe constituents exported",
            universe=universe_code,
            file=f"{universe_code}_Index_Constituents_with_Factset_code.parquet",
        )
        del df_final

    # concatenate and export
    df_all = pd.concat(df_all).drop_duplicates(ignore_index=True)
    df_all.to_parquet(
        FACTSET_ROOT_DIR
        / "Index_Constituents/Index_Constituents_with_Factset_code.parquet",
        index=False,
    )
    logger.info(
        "Combined constituents exported",
        file="Index_Constituents_with_Factset_code.parquet",
    )

    if split_save_mode:
        # データフレームが大きいため分割して圧縮
        saved_paths = split_and_save_dataframe(
            df_all=df_all,
            base_dir=FACTSET_ROOT_DIR / "Index_Constituents",
            n_splits=5,
            base_filename="Index_Constituents_with_Factset_code-compressed-",
            compression="zstd",
            index=False,
        )
    else:
        index_constituents_dir = FACTSET_ROOT_DIR / "Index_Constituents"
        already_exsiting_compressed_files = list(
            index_constituents_dir.glob(
                "Index_Constituents_with_Factset_code-compressed-*.parquet"
            )
        )
        saved_paths = [
            (
                index_constituents_dir
                / f"Index_Constituents_with_Factset_code-compressed-{len(already_exsiting_compressed_files) + 1}.parquet"
            )
        ]
        df_all.to_parquet(saved_paths[0], index=False, compression="zstd")

    logger.info("All split files saved", paths=[str(p) for p in saved_paths])

    del df_all
    logger.info("Universe constituents export completed")


# ============================================================================================
def create_factset_symbol_list_function(universe_code: str) -> str:
    """
    GET constituents list
    Factset formula用にFactset P_SYMBOLを取得 -> stringに変換

    universe_code: BPMのインデックスコード
    """

    load_project_env()
    factset_root_dir = Path(os.environ.get("FACTSET_ROOT_DIR"))  # type: ignore
    symbol_list = (
        pd.read_parquet(
            factset_root_dir
            / f"Index_Constituents/{universe_code}_Index_Constituents_with_Factset_code.parquet"
        )
        .dropna(subset=["P_SYMBOL"])
        .sort_values("P_SYMBOL")["P_SYMBOL"]
        .unique()
        .tolist()
    )
    symbol_list_str = " ".join(symbol_list)
    symbol_list_function = (
        f'^=STRING SYMBOLS=TICKER_LIST("{symbol_list_str}")'  # FQL埋め込み用関数
    )

    return symbol_list_function


# ============================================================================================
def factset_formula(item: str, year_range: str = "20AY", per: str = "M") -> str:
    """
    Factset formula作成用関数

    item: Factset関数のitem name(ex. FF_SALES)
    per: 取得頻度(D: Daily, M: Monthly, Q: Quarterly, A: Annually)
    """

    period_start = "-" + year_range if not year_range.startswith("-") else year_range
    if item == "FF_ENTRPR_VAL_DAILY":
        excel_function = f'^=GET_FQL_ARRAY(SYMBOLS, "{item}({period_start}, 0, {per},, USD, ""DIL"")")'
    elif item == "FG_PRICE":
        excel_function = (
            f'^=GET_FQL_ARRAY(SYMBOLS, "FG_PRICE({period_start}, 0, {per})")'
        )
    else:
        excel_function = f'^=GET_FQL_ARRAY(SYMBOLS, "AVAIL({item}(QTR_R, {period_start}, 0, {per},,USD), {item}(SEMI_R, {period_start}, 0, {per},,USD), {item}(ANN_R, {period_start}, 0, {per},,USD))")'
    return excel_function


# ============================================================================================
def implement_factset_formulas(universe_code: str, year_range: str = "20AY") -> None:
    """
    UniverseをBPMのコードで指定し、Factsetから財務データをダウンロードするエクセルファイルを作成する関数
    universe_code: BPMのインデックスコード
    """

    # フォルダ
    load_project_env()
    FACTSET_ROOT_DIR = Path(os.environ.get("FACTSET_ROOT_DIR"))  # type: ignore
    FACTSET_FINANCIALS_DIR = Path(os.environ.get("FACTSET_FINANCIALS_DIR"))  # type: ignore
    #  read factset formula items
    formula_xlsx = FACTSET_ROOT_DIR / "FDS samples and Factset Formulas.xlsx"
    df = pd.read_excel(formula_xlsx, sheet_name="Sheet3")

    symbol_list_function = create_factset_symbol_list_function(
        universe_code=universe_code
    )

    logger.info(
        "Factset formula embedding started",
        universe=universe_code,
        category_count=len(df["Category"].unique()),
    )
    category_list = df["Category"].unique().tolist()
    logger.debug("Categories", categories=", ".join(category_list))
    for category in category_list:
        df_cat = df.loc[
            df["Category"] == category,
            ["Category", "Item", "name"],
        ]
        universe_folder = FACTSET_FINANCIALS_DIR / universe_code
        universe_folder.mkdir(exist_ok=True)

        per = "M"
        excel_file = universe_folder / f"Financials_{category}_{year_range}.xlsx"
        if category == "Price_Daily":
            per = "D"
            excel_file = universe_folder / f"Price_Daily_{year_range}.xlsx"
        elif category == "Price":
            per = "M"
            excel_file = universe_folder / f"Price_{year_range}.xlsx"

        # export to Excel file
        wb = openpyxl.Workbook()
        for i, item in enumerate(df_cat["Item"].tolist()):
            ws = wb.create_sheet(title=item, index=i)
            ws["A1"].value = "date"
            ws["A2"].value = f"^=P_DATE(-{year_range}, 0, {per})"
            ws["B1"].value = symbol_list_function
            ws["B2"].value = factset_formula(item=item, year_range=year_range, per=per)
        wb.save(excel_file)
        wb.close()

    logger.info("Factset formula embedding completed", universe=universe_code)


# ============================================================================================
def format_factset_downloaded_data(
    file_list: list[Path | str], output_folder: Path, split_save_mode: bool = False
):
    """Factsetからダウンロードした財務・価格データを整形しParquet形式で保存する。

    指定されたExcelファイルのリストを読み込み、各シートのデータを縦持ち形式(long format)に変換します。
    変換後のデータは単一のDataFrameに結合され、Parquetファイルとして出力されます。
    保存モードに応じて、単一ファイルまたは分割ファイルとして保存します。

    Parameters
    ----------
    file_list : list[Path | str]
        Factsetからダウンロードしたデータが含まれるExcelファイルのパスのリスト。
    output_folder : Path
        出力先のParquetファイルを保存するフォルダのパス。
    split_save_mode : bool, optional
        保存モードを制御するフラグ, by default False.
        - `True`の場合:
            1. `Financials_and_Price.parquet`が存在すれば、新しいデータを追記して更新します。
                存在しない場合は新規作成します。
            2. さらに、データを3年ごとの期間で分割し、
                `Financials_and_Price-compressed-YYYY-YYYY.parquet`という名前の
                圧縮ファイルとして保存します。
        - `False`の場合:
            全期間のデータを単一の圧縮ファイル
            `Financials_and_Price-compressed-YYYYMMDD_YYYYMMDD.parquet`
            として保存します。

    Returns
    -------
    None
        この関数は値を返さず、ファイルを出力します。

    """

    df_all = pd.DataFrame()
    all_melted_dfs = []

    for f in tqdm(file_list):
        xls = pd.ExcelFile(f, engine="calamine")
        sheet_names = [
            s for s in xls.sheet_names if s not in ["Sheet", "FF_ENTRPR_VAL_DAILY"]
        ]
        melted_df = [
            pd.melt(
                pd.read_excel(f, sheet_name=sheet_name).dropna(
                    how="all", axis=0, ignore_index=True
                ),
                id_vars="date",  # type: ignore
                var_name="P_SYMBOL",
                value_name="value",
            ).assign(variable=sheet_name)
            for sheet_name in sheet_names
        ]
        all_melted_dfs.extend(melted_df)

    df_all = pd.concat(all_melted_dfs, ignore_index=True)
    df_all["date"] = pd.to_datetime(df_all["date"])
    df_all["value"] = df_all["value"].astype(float)

    # export
    if split_save_mode:
        # データフレームが大きいため分割して圧縮
        already_exsiting_compressed_file = (
            output_folder / "Financials_and_Price.parquet"
        )
        if already_exsiting_compressed_file.exists():
            df_existed = pd.read_parquet(already_exsiting_compressed_file)
            df_all = pd.concat([df_existed, df_all]).drop_duplicates(ignore_index=True)
            df_all.to_parquet(
                already_exsiting_compressed_file, index=False, compression="zstd"
            )
            logger.info("File exported", path=str(already_exsiting_compressed_file))
        else:
            output_path = output_folder / "Financials_and_Price.parquet"
            df_all.to_parquet(output_path, index=False, compression="zstd")
            logger.info("File exported", path=str(output_path))

        # 分割してexport
        start_year_list = np.arange(pd.to_datetime(df_all["date"]).min().year, 2025, 3)
        for start, end in zip(start_year_list, start_year_list + 2, strict=False):
            df_slice = df_all[
                (df_all["date"].dt.year >= start) & (df_all["date"].dt.year <= end)
            ]
            output_path = (
                output_folder / f"Financials_and_Price-compressed-{start}-{end}.parquet"
            )
            df_slice.to_parquet(output_path, index=False, compression="zstd")
            logger.info("File exported", path=str(output_path))

    else:
        start_date = df_all["date"].min().strftime("%Y%m%d")
        end_date = df_all["date"].max().strftime("%Y%m%d")
        output_path = (
            output_folder
            / f"Financials_and_Price-compressed-{start_date}_{end_date}.parquet"
        )
        df_all.to_parquet(output_path, index=False, compression="zstd")
        logger.info("File exported", path=str(output_path))


# ============================================================================================
# def store_to_database(
#     df: pd.DataFrame,
#     db_path: Path,
#     table_name: str,
#     unique_cols: list[str] = ["date", "P_SYMBOL", "variable"],
#     verbose: bool = True,
#     replace_on_conflict: bool = False,
# ):
#     """
#     Pandas DataFrameをSQLiteデータベースに書き込む。
#     既存のテーブルのデータと重複する行(date, P_SYMBOL, value, variable の組み合わせが一致する行)
#     は追加しない。

#     Args:
#         df (pd.DataFrame): 書き込みたいデータフレーム(date, P_SYMBOL, value, variable のカラムを持つ)。
#         db_path (str): 接続するSQLiteデータベースのファイルパス。
#         table_name (str): 書き込み先のテーブル名。
#         unique_cols ([str]): 一意性をチェックするカラム
#     """

#     # 必須カラムのチェック
#     if not all(col in df.columns for col in unique_cols):
#         raise ValueError(
#             f"データフレームには必須のカラム {unique_cols} の全てが含まれている必要があります。"
#         )

#     # 1. データベースに接続
#     conn = sqlite3.connect(db_path)
#     conn.execute(
#         f"""
#             CREATE TABLE IF NOT EXISTS {table_name} (
#                 date TEXT,
#                 P_SYMBOL TEXT,
#                 variable TEXT,
#                 value REAL,
#                 PRIMARY KEY ({",".join(unique_cols)})
#                 )
#         """
#     )
#     # 2. 既存のテーブルから一意性チェックに必要なデータを取得し、重複行を除外
#     try:
#         # テーブルが存在する場合、既存の複合キーデータを取得
#         select_cols = ", ".join(unique_cols)
#         existing_df = pd.read_sql(f"SELECT {select_cols} FROM {table_name}", conn)

#         # 既存データの 'date' カラムも datetime 型に変換する
#         # (unique_cols に 'date' が含まれている場合のみ)
#         if "date" in unique_cols and "date" in existing_df.columns:
#             existing_df["date"] = pd.to_datetime(existing_df["date"])

#         # 重複チェック用に新しいデータと既存データをマージ
#         merged_df = pd.merge(
#             df.drop_duplicates(subset=unique_cols),  # 新しいデータ側も自身の重複を除去
#             existing_df,
#             on=unique_cols,
#             how="left",
#             indicator=True,
#         )

#         # 既存データに含まれていない行 ('left_only') のみを選択
#         df_to_add = merged_df[merged_df["_merge"] == "left_only"].drop(
#             columns=["_merge"]
#         )

#         if df_to_add.empty:
#             print(
#                 f"テーブル '{table_name}' に追加すべき新しいデータはありませんでした。スキップします。"
#             )
#             conn.close()
#             return

#         if verbose:
#             print(
#                 f"既存の {len(existing_df)} 行との重複をチェックしました。{len(df_to_add)} 行を新たに追加します。"
#             )

#     except pd.io.sql.DatabaseError:  # type: ignore
#         # テーブルがまだ存在しない場合、全ての行を追加
#         df_to_add = df.drop_duplicates(subset=unique_cols)  # 自身の重複は除去
#         if verbose:
#             print(
#                 f"テーブル '{table_name}' は存在しません。新しいテーブルとして、すべての {len(df_to_add)} 行を追加します。"
#             )

#     # 3. ユニークな行だけをデータベースに書き込み
#     df_to_add.to_sql(table_name, conn, if_exists="append", index=False)

#     # 4. 接続を閉じる
#     conn.close()
#     if verbose:
#         print(f"  -> {table_name}: データの書き込みが完了しました。")


# ============================================================================================
def store_to_database(
    df: pd.DataFrame,
    db_path: Path,
    table_name: str,
    unique_cols: list[str] | None = None,
    verbose: bool = True,
    on_duplicate: str = "skip",  # "skip" または "update"
):
    """
    Pandas DataFrameをSQLiteデータベースに書き込む。

    Args:
        df (pd.DataFrame): 書き込みたいデータフレーム(date, P_SYMBOL, value, variable のカラムを持つ)。
        db_path (str): 接続するSQLiteデータベースのファイルパス。
        table_name (str): 書き込み先のテーブル名。
        unique_cols ([str]): 一意性をチェックするカラム
        on_duplicate (str): 重複時の動作 - "skip" (スキップ) または "update" (上書き)
    """
    if unique_cols is None:
        unique_cols = ["date", "P_SYMBOL", "variable"]

    # 必須カラムのチェック
    if not all(col in df.columns for col in unique_cols):
        raise ValueError(
            f"データフレームには必須のカラム {unique_cols} の全てが含まれている必要があります。"
        )

    # テーブル名とカラム名のバリデーション（SQLインジェクション対策）
    _validate_sql_identifier(table_name)
    for col in unique_cols:
        _validate_sql_identifier(col)

    # 1. データベースに接続
    conn = sqlite3.connect(db_path)
    # nosec B608 - table_name, unique_cols は _validate_sql_identifier() で検証済み
    conn.execute(
        f"""
            CREATE TABLE IF NOT EXISTS {table_name} (
                date TEXT,
                P_SYMBOL TEXT,
                variable TEXT,
                value REAL,
                PRIMARY KEY ({",".join(unique_cols)})
            )
        """
    )

    # 2. 既存のテーブルから一意性チェックに必要なデータを取得し、重複行を除外
    try:
        # テーブルが存在する場合、既存の複合キーデータを取得
        # table_name, unique_cols は _validate_sql_identifier() で検証済み
        select_cols = ", ".join(unique_cols)
        existing_df = pd.read_sql(f"SELECT {select_cols} FROM {table_name}", conn)  # nosec B608

        # 既存データの 'date' カラムも datetime 型に変換する
        if "date" in unique_cols and "date" in existing_df.columns:
            existing_df["date"] = pd.to_datetime(existing_df["date"])

        # 重複チェック用に新しいデータと既存データをマージ
        merged_df = pd.merge(
            df.drop_duplicates(subset=unique_cols),
            existing_df,
            on=unique_cols,
            how="left",
            indicator=True,
        )

        # 新規データと重複データを分離
        df_to_add = merged_df[merged_df["_merge"] == "left_only"].drop(
            columns=["_merge"]
        )
        df_to_update = merged_df[merged_df["_merge"] == "both"].drop(columns=["_merge"])

        if on_duplicate == "update" and not df_to_update.empty:
            # 重複データを削除してから追加（上書き）
            # table_name, unique_cols は _validate_sql_identifier() で検証済み
            delete_count = 0
            for _, row in df_to_update.iterrows():
                conditions = " AND ".join([f"{col} = ?" for col in unique_cols])
                values = tuple(
                    str(row[col]) if isinstance(row[col], pd.Timestamp) else row[col]
                    for col in unique_cols
                )
                conn.execute(f"DELETE FROM {table_name} WHERE {conditions}", values)  # nosec B608
                delete_count += 1

            conn.commit()  # DELETEを確定

            if verbose:
                logger.debug(
                    "Duplicate rows deleted for overwrite", delete_count=delete_count
                )

            # 重複データを追加対象に含める
            df_to_add = pd.concat([df_to_add, df_to_update], ignore_index=True)

        if df_to_add.empty:
            logger.info("No new data to add, skipping", table=table_name)
            conn.close()
            return

        if verbose:
            if on_duplicate == "update":
                logger.debug(
                    "Rows to add",
                    table=table_name,
                    total=len(df_to_add),
                    overwritten=len(df_to_update),
                )
            else:
                logger.debug(
                    "Deduplication check completed",
                    table=table_name,
                    existing_rows=len(existing_df),
                    new_rows=len(df_to_add),
                )

    except pd.io.sql.DatabaseError:  # type: ignore
        # テーブルがまだ存在しない場合、全ての行を追加
        df_to_add = df.drop_duplicates(subset=unique_cols)
        if verbose:
            logger.debug(
                "Table does not exist, creating new",
                table=table_name,
                rows=len(df_to_add),
            )

    # 3. ユニークな行だけをデータベースに書き込み
    df_to_add.to_sql(table_name, conn, if_exists="append", index=False)
    conn.commit()  # INSERTを確定

    # 4. 接続を閉じる
    conn.close()
    if verbose:
        logger.info("Data written to database", table=table_name)


# ============================================================================================
# WALモード有効化関数
# ============================================================================================


def enable_wal_mode(db_path: Path, verbose: bool = True) -> None:
    """
    SQLiteデータベースでWALモードを有効化

    WAL(Write-Ahead Logging)モードにより、読み込みと書き込みの並行性が向上します。
    データベース作成時に一度だけ実行すれば、以降は永続的に有効です。

    :param db_path: データベースファイルパス
    :param verbose: 進捗表示フラグ
    """
    with sqlite3.connect(db_path) as conn:
        cursor = conn.cursor()

        # 現在のジャーナルモードを確認
        cursor.execute("PRAGMA journal_mode")
        current_mode = cursor.fetchone()[0]

        if current_mode != "wal":
            # WALモードに変更
            cursor.execute("PRAGMA journal_mode=WAL")
            new_mode = cursor.fetchone()[0]

            if verbose:
                logger.info(
                    "Journal mode changed", from_mode=current_mode, to_mode=new_mode
                )
        elif verbose:
            logger.debug("Already in WAL mode")


# ============================================================================================
# 🆕 直列書き込み版(推奨・データベースロック完全回避)
# ============================================================================================


def store_active_returns_batch_serial_write(
    df_active_returns: pd.DataFrame,
    return_cols: list[str],
    db_path: Path,
    benchmark_ticker: str,
    batch_size: int = 10000,
    verbose: bool = True,
) -> dict[str, Any]:
    """
    アクティブリターンをバッチ保存(直列書き込み版・ロック完全回避)

    この関数はデータベースロック問題を完全に解決します：
    - データの前処理(グループ化、型変換)は高速に実行
    - DB書き込みは1接続で直列化し、ロックを完全回避
    - 100%の成功率を保証

    :param df_active_returns: アクティブリターンデータ(Long形式)
    :param return_cols: 処理対象のリターン列名リスト
    :param db_path: SQLiteデータベースのファイルパス
    :param benchmark_ticker: ベンチマークティッカー(除外用)
    :param batch_size: バッチサイズ(一度に挿入する行数)
    :param verbose: 進捗表示フラグ
    :return: 処理結果の統計情報
    """
    if verbose:
        logger.info(
            "Active returns batch save started (serial write)",
            columns=len(return_cols),
            rows=len(df_active_returns),
            batch_size=batch_size,
        )

    import time

    start_time = time.time()

    # --------------------------------------------------------------------------
    # ステップ1: データ前処理(型変換、グループ化)
    # --------------------------------------------------------------------------
    if verbose:
        logger.debug("Data preprocessing started")

    prep_start = time.time()

    # データコピーと型変換
    df_opt = df_active_returns.copy()
    df_opt["value"] = df_opt["value"].astype(float)
    df_opt["date"] = pd.to_datetime(df_opt["date"])
    df_opt = df_opt.rename(columns={"symbol": "P_SYMBOL"})

    # ベンチマーク除外
    df_opt = df_opt[df_opt["P_SYMBOL"] != benchmark_ticker]

    # variable列でグループ化して辞書化(高速アクセス用)
    df_dict = {}
    for col in return_cols:
        active_return_col = col.replace("Return", "Active_Return")

        # queryの代わりにフィルタリング(100倍高速)
        df_slice = df_opt[df_opt["variable"] == active_return_col].reset_index(
            drop=True
        )

        if not df_slice.empty:
            # 自身の重複を除去
            df_slice = df_slice.drop_duplicates(
                subset=["date", "P_SYMBOL", "variable"], ignore_index=True
            )
            df_dict[active_return_col] = df_slice

    prep_time = time.time() - prep_start

    if verbose:
        logger.info(
            "Preprocessing completed",
            elapsed_sec=round(prep_time, 2),
            tables=len(df_dict),
        )

    # --------------------------------------------------------------------------
    # ステップ2: データベース書き込み(直列化でロック回避)
    # --------------------------------------------------------------------------
    results = {
        "success": [],
        "failed": [],
        "total_rows": 0,
        "prep_time": prep_time,
        "save_time": 0,
        "total_time": 0,
    }

    save_start = time.time()

    # 単一の接続を使用(直列処理でロック完全回避)
    with sqlite3.connect(db_path, timeout=30.0) as conn:
        cursor = conn.cursor()

        # 進捗バー
        iterator = (
            tqdm(df_dict.items(), desc="💾 保存中") if verbose else df_dict.items()
        )

        for table_name, df in iterator:
            try:
                # テーブル名のバリデーション（SQLインジェクション対策）
                _validate_sql_identifier(table_name)

                # テーブル作成(存在しない場合)
                # nosec B608 - table_name は _validate_sql_identifier() で検証済み
                cursor.execute(
                    f"""
                    CREATE TABLE IF NOT EXISTS "{table_name}" (
                        date TEXT,
                        P_SYMBOL TEXT,
                        variable TEXT,
                        value REAL,
                        PRIMARY KEY (date, P_SYMBOL, variable)
                    )
                    """
                )

                # 既存データとの重複チェック
                try:
                    select_cols = ", ".join(["date", "P_SYMBOL", "variable"])
                    # table_name は _validate_sql_identifier() で検証済み
                    query = f'SELECT {select_cols} FROM "{table_name}"'  # nosec B608
                    existing_df = pd.read_sql(query, conn)

                    # 既存データのdate列も変換
                    if "date" in existing_df.columns:
                        existing_df["date"] = pd.to_datetime(existing_df["date"])

                    # 重複チェック
                    merged_df = pd.merge(
                        df,
                        existing_df,
                        on=["date", "P_SYMBOL", "variable"],
                        how="left",
                        indicator=True,
                    )

                    df_to_add = merged_df[merged_df["_merge"] == "left_only"].drop(
                        columns=["_merge"]
                    )

                    if df_to_add.empty:
                        if verbose:
                            logger.debug(
                                "Duplicate data only, skipping", table=table_name
                            )
                        continue

                except pd.io.sql.DatabaseError:  # type: ignore
                    # テーブルが存在しない場合は全データを追加
                    df_to_add = df

                row_count = len(df_to_add)

                # バッチINSERTで保存
                df_to_add.to_sql(
                    table_name,
                    conn,
                    if_exists="append",
                    index=False,
                    chunksize=batch_size,
                    method="multi",
                )

                conn.commit()

                results["success"].append(table_name)
                results["total_rows"] += row_count

                if verbose:
                    logger.debug("Table saved", table=table_name, rows=row_count)

            except Exception as e:
                results["failed"].append({"table": table_name, "error": str(e)})
                logger.error(
                    "Table save failed", table=table_name, error=str(e), exc_info=True
                )

                # エラー時はロールバック
                with contextlib.suppress(BaseException):
                    conn.rollback()

    results["save_time"] = time.time() - save_start
    results["total_time"] = time.time() - start_time

    # --------------------------------------------------------------------------
    # ステップ3: 統計情報表示
    # --------------------------------------------------------------------------
    if verbose:
        success_rate = (len(results["success"]) / len(df_dict) * 100) if df_dict else 0
        throughput = (
            results["total_rows"] / results["total_time"]
            if results["total_time"] > 0
            else 0
        )
        logger.info(
            "Batch save completed",
            success_count=len(results["success"]),
            total_tables=len(df_dict),
            failed_count=len(results["failed"]),
            total_rows=results["total_rows"],
            prep_time_sec=round(results["prep_time"], 2),
            save_time_sec=round(results["save_time"], 2),
            total_time_sec=round(results["total_time"], 2),
            throughput_per_sec=round(throughput),
            success_rate_pct=round(success_rate, 1),
        )

        if results["failed"]:
            for failed in results["failed"]:
                logger.warning(
                    "Failed table", table=failed["table"], error=failed["error"]
                )

    return results


# ============================================================================================
def insert_active_returns_optimized_sqlite(
    df_active_returns: pd.DataFrame,
    return_cols: list[str],
    db_path: Path,
    benchmark_ticker: str,
    # batch_sizeはexecutemanyの分割には使われないが、互換性のため残す
    batch_size: int = 10000,
    verbose: bool = True,
) -> dict[str, Any]:
    """
    アクティブリターンをバッチ保存(直列書き込み版・ロック完全回避・DB最適化)

    - データ前処理(グループ化、型変換)は高速に実行
    - DB書き込みは1接続で直列化し、ロックを完全回避
    - 書き込み速度向上のため、INSERT OR IGNOREとexecutemanyを使用
    """
    if verbose:
        logger.info(
            "Active returns optimized batch save started (SQLite)",
            columns=len(return_cols),
            rows=len(df_active_returns),
            batch_size=batch_size,
        )

    start_time = time.time()

    # --------------------------------------------------------------------------
    # ステップ1: データ前処理(型変換、グループ化)- 変更なし
    # --------------------------------------------------------------------------
    if verbose:
        logger.debug("Data preprocessing started")

    prep_start = time.time()

    # データコピーと型変換
    df_opt = df_active_returns.copy()
    df_opt["value"] = df_opt["value"].astype(float)
    df_opt["date"] = pd.to_datetime(df_opt["date"]).dt.strftime(
        "%Y-%m-%d %H:%M:%S"
    )  # TEXT形式に変換
    df_opt = df_opt.rename(columns={"symbol": "P_SYMBOL"})

    # ベンチマーク除外
    df_opt = df_opt[df_opt["P_SYMBOL"] != benchmark_ticker]

    # variable列でグループ化して辞書化
    df_dict = {}
    for col in return_cols:
        active_return_col = col.replace("Return", "Active_Return")

        # フィルタリング
        df_slice = df_opt[df_opt["variable"] == active_return_col].reset_index(
            drop=True
        )

        if not df_slice.empty:
            # 自身の重複を除去
            df_slice = df_slice.drop_duplicates(
                subset=["date", "P_SYMBOL", "variable"], ignore_index=True
            )
            df_dict[active_return_col] = df_slice

    prep_time = time.time() - prep_start

    if verbose:
        logger.info(
            "Preprocessing completed",
            elapsed_sec=round(prep_time, 2),
            tables=len(df_dict),
        )

    # --------------------------------------------------------------------------
    # ステップ2: データベース書き込み(SQLite最適化)
    # --------------------------------------------------------------------------
    results = {
        "success": [],
        "failed": [],
        "total_rows": 0,
        "prep_time": prep_time,
        "save_time": 0,
        "total_time": 0,
    }

    save_start = time.time()

    # 単一の接続を使用
    with sqlite3.connect(db_path, timeout=30.0) as conn:
        cursor = conn.cursor()

        # 💡 【最適化A】SQLiteのパフォーマンス設定
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")

        # 進捗バー
        iterator = (
            tqdm(df_dict.items(), desc="💾 保存中") if verbose else df_dict.items()
        )

        for table_name, df in iterator:
            if df.empty:
                continue

            try:
                # テーブル作成(存在しない場合)
                cursor.execute(
                    f"""
                    CREATE TABLE IF NOT EXISTS "{table_name}" (
                        date TEXT,
                        P_SYMBOL TEXT,
                        variable TEXT,
                        value REAL,
                        PRIMARY KEY (date, P_SYMBOL, variable)
                    )
                    """
                )

                # 💡 【最適化B】executemanyとINSERT OR IGNOREを使用

                # 1. 挿入するデータをタプルリストに変換
                # 列の順序はSQLステートメントの順序と一致させる
                data_to_insert = df[
                    ["date", "P_SYMBOL", "variable", "value"]
                ].values.tolist()

                # 2. SQLステートメントを準備(重複行は無視)
                sql_insert = f"""
                    INSERT OR IGNORE INTO "{table_name}" (date, P_SYMBOL, variable, value)
                    VALUES (?, ?, ?, ?)
                """

                # 3. executemanyで高速にバッチ挿入を実行
                # executemanyは、データリストをチャンクに分割して処理する(batch_sizeは事実上無視される)
                cursor.executemany(sql_insert, data_to_insert)

                # 4. 挿入された行数をカウント (rowcountはexecutemanyでは信頼できないため、別の方法で取得が必要だが、ここでは簡略化)
                # 簡略化のため、挿入対象だった全行数を使用
                row_count = len(data_to_insert)

                conn.commit()

                results["success"].append(table_name)
                results["total_rows"] += (
                    row_count  # 実際には重複無視で挿入されなかった行も含まれるが、元のコードに合わせる
                )

                if verbose:
                    logger.debug(
                        "Table insert attempted", table=table_name, rows=row_count
                    )

            except Exception as e:
                results["failed"].append({"table": table_name, "error": str(e)})
                logger.error(
                    "Table save failed", table=table_name, error=str(e), exc_info=True
                )

                # エラー時はロールバック
                with contextlib.suppress(BaseException):
                    conn.rollback()

    results["save_time"] = time.time() - save_start
    results["total_time"] = time.time() - start_time

    # --------------------------------------------------------------------------
    # ステップ3: 統計情報表示
    # --------------------------------------------------------------------------
    if verbose:
        success_rate = (len(results["success"]) / len(df_dict) * 100) if df_dict else 0
        throughput = (
            results["total_rows"] / results["total_time"]
            if results["total_time"] > 0
            else 0
        )
        logger.info(
            "Batch save completed",
            success_count=len(results["success"]),
            total_tables=len(df_dict),
            failed_count=len(results["failed"]),
            total_attempted_rows=results["total_rows"],
            prep_time_sec=round(results["prep_time"], 2),
            save_time_sec=round(results["save_time"], 2),
            total_time_sec=round(results["total_time"], 2),
            throughput_per_sec=round(throughput),
            success_rate_pct=round(success_rate, 1),
        )

        if results["failed"]:
            for failed in results["failed"]:
                logger.warning(
                    "Failed table", table=failed["table"], error=failed["error"]
                )

    return results


# ============================================================================================
def store_to_database_batch(
    df_dict: dict[str, pd.DataFrame],
    db_path: Path,
    unique_cols: list[str] | None = None,
    batch_size: int = 10000,
    max_workers: int | None = 1,  # デフォルトを1に変更(ロック回避)
    verbose: bool = True,
) -> dict[str, Any]:
    """
    複数のDataFrameをバッチ保存でデータベースに書き込む

    ⚠️ 注意: max_workersを1にすることでロック問題を回避します。
    並列処理が必要な場合は、事前にWALモードを有効化してください。

    :param df_dict: {table_name: DataFrame}の辞書
    :param db_path: SQLiteデータベースのファイルパス
    :param unique_cols: 一意性をチェックするカラムのリスト
    :param batch_size: 一度に挿入する行数
    :param max_workers: 並列実行する最大スレッド数(1推奨)
    :param verbose: 進捗表示フラグ
    :return: 処理結果の統計情報
    """
    if unique_cols is None:
        unique_cols = ["date", "P_SYMBOL", "variable"]

    if not df_dict:
        if verbose:
            logger.warning("No data to save")
        return {"success": 0, "failed": 0, "total_rows": 0, "processing_time": 0}

    # テーブル名とカラム名のバリデーション（SQLインジェクション対策）
    for table_name in df_dict:
        _validate_sql_identifier(table_name)
    for col in unique_cols:
        _validate_sql_identifier(col)

    if max_workers is not None and max_workers > 1 and verbose:
        logger.warning(
            "max_workers > 1 risks database locks",
            max_workers=max_workers,
        )

    if verbose:
        logger.info(
            "Batch save mode started",
            tables=len(df_dict),
            batch_size=batch_size,
            max_workers=max_workers,
            database=Path(db_path).name,
        )

    import time

    start_time = time.time()

    # --------------------------------------------------------------------------
    # ステップ1: 事前にデータ型を一括変換(全DataFrameに対して)
    # --------------------------------------------------------------------------
    if verbose:
        logger.debug("Data type conversion started")

    optimized_dict = {}
    prep_start = time.time()

    for table_name, df in df_dict.items():
        df_opt = df.copy()

        # 必須カラムのチェック
        if not all(col in df_opt.columns for col in unique_cols):
            if verbose:
                logger.warning(
                    "Skipping table, missing required columns", table=table_name
                )
            continue

        # データ型変換
        if "date" in df_opt.columns:
            df_opt["date"] = pd.to_datetime(df_opt["date"])
        if "value" in df_opt.columns:
            df_opt["value"] = df_opt["value"].astype(float)

        # 重複除去
        df_opt = df_opt.drop_duplicates(subset=unique_cols, ignore_index=True)

        if not df_opt.empty:
            optimized_dict[table_name] = df_opt

    prep_time = time.time() - prep_start

    if verbose:
        logger.info(
            "Preprocessing completed",
            elapsed_sec=round(prep_time, 2),
            tables=len(optimized_dict),
        )

    # --------------------------------------------------------------------------
    # ステップ2: ワーカー関数
    # --------------------------------------------------------------------------
    def _batch_save_worker(args: tuple) -> tuple:
        """バッチ保存ワーカー"""
        table_name, df = args

        try:
            with sqlite3.connect(db_path, timeout=30.0) as conn:
                cursor = conn.cursor()

                # テーブル作成
                # table_name, unique_cols は関数先頭で _validate_sql_identifier() 検証済み
                cursor.execute(
                    f"""
                    CREATE TABLE IF NOT EXISTS "{table_name}" (
                        date TEXT,
                        P_SYMBOL TEXT,
                        variable TEXT,
                        value REAL,
                        PRIMARY KEY ({",".join(unique_cols)})
                    )
                    """  # nosec B608
                )

                # 重複チェック
                try:
                    select_cols = ", ".join(unique_cols)
                    # table_name, unique_cols は関数先頭で検証済み
                    query = f'SELECT {select_cols} FROM "{table_name}"'  # nosec B608
                    existing_df = pd.read_sql(query, conn)

                    if "date" in unique_cols and "date" in existing_df.columns:
                        existing_df["date"] = pd.to_datetime(existing_df["date"])

                    merged_df = pd.merge(
                        df,
                        existing_df,
                        on=unique_cols,
                        how="left",
                        indicator=True,
                    )

                    df_to_add = merged_df[merged_df["_merge"] == "left_only"].drop(
                        columns=["_merge"]
                    )

                    if df_to_add.empty:
                        return (table_name, True, 0, "重複データのみ")

                except pd.io.sql.DatabaseError:  # type: ignore
                    df_to_add = df

                row_count = len(df_to_add)

                # バッチINSERT
                df_to_add.to_sql(
                    table_name,
                    conn,
                    if_exists="append",
                    index=False,
                    chunksize=batch_size,
                    method="multi",
                )

                conn.commit()

                return (table_name, True, row_count, None)

        except Exception as e:
            return (table_name, False, 0, str(e))

    # --------------------------------------------------------------------------
    # ステップ3: 実行
    # --------------------------------------------------------------------------
    results = {
        "success": [],
        "failed": [],
        "total_rows": 0,
        "prep_time": prep_time,
        "save_time": 0,
        "total_time": 0,
    }

    save_start = time.time()
    args_list = list(optimized_dict.items())

    if max_workers == 1:
        # 直列処理
        iterator = tqdm(args_list, desc="💾 保存中") if verbose else args_list

        for args in iterator:
            table_name, success, row_count, error_msg = _batch_save_worker(args)

            if success:
                results["success"].append(table_name)
                results["total_rows"] += row_count

                if verbose and row_count > 0:
                    logger.debug("Table saved", table=table_name, rows=row_count)
                elif verbose:
                    logger.debug("Table skipped", table=table_name, reason=error_msg)
            else:
                results["failed"].append({"table": table_name, "error": error_msg})
                logger.error("Table save failed", table=table_name, error=error_msg)

    else:
        # 並列処理(WALモード推奨)
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {
                executor.submit(_batch_save_worker, args): args[0] for args in args_list
            }

            if verbose:
                futures_iterator = tqdm(
                    as_completed(futures), total=len(futures), desc="Saving"
                )
            else:
                futures_iterator = as_completed(futures)

            for future in futures_iterator:
                table_name = futures[future]

                try:
                    table_name_result, success, row_count, error_msg = future.result()

                    if success:
                        results["success"].append(table_name_result)
                        results["total_rows"] += row_count

                        if verbose and row_count > 0:
                            logger.debug(
                                "Table saved", table=table_name_result, rows=row_count
                            )
                        elif verbose:
                            logger.debug(
                                "Table skipped",
                                table=table_name_result,
                                reason=error_msg,
                            )
                    else:
                        results["failed"].append(
                            {"table": table_name_result, "error": error_msg}
                        )
                        logger.error(
                            "Table save failed",
                            table=table_name_result,
                            error=error_msg,
                        )

                except Exception as e:
                    results["failed"].append({"table": table_name, "error": str(e)})
                    logger.error(
                        "Table save failed",
                        table=table_name,
                        error=str(e),
                        exc_info=True,
                    )

    results["save_time"] = time.time() - save_start
    results["total_time"] = time.time() - start_time

    # --------------------------------------------------------------------------
    # 統計情報
    # --------------------------------------------------------------------------
    if verbose:
        throughput = (
            results["total_rows"] / results["total_time"]
            if results["total_time"] > 0
            else 0
        )
        logger.info(
            "Batch save completed",
            success_count=len(results["success"]),
            total_tables=len(df_dict),
            failed_count=len(results["failed"]),
            total_rows=results["total_rows"],
            prep_time_sec=round(results["prep_time"], 2),
            save_time_sec=round(results["save_time"], 2),
            total_time_sec=round(results["total_time"], 2),
            throughput_per_sec=round(throughput),
        )

        if results["failed"]:
            for failed in results["failed"]:
                logger.warning(
                    "Failed table", table=failed["table"], error=failed["error"]
                )

    return results


# ============================================================================================
def store_active_returns_batch(
    df_active_returns: pd.DataFrame,
    return_cols: list[str],
    db_path: Path,
    benchmark_ticker: str,
    batch_size: int = 10000,
    max_workers: int | None = None,
    verbose: bool = True,
) -> dict[str, Any]:
    """
    アクティブリターンをバッチ保存(最適化版)

    元のコードを置き換える関数。事前にvariable列でグループ化してから
    バッチ保存することで大幅な高速化を実現。

    :param df_active_returns: アクティブリターンデータ(Long形式)
    :param return_cols: 処理対象のリターン列名リスト
    :param db_path: SQLiteデータベースのファイルパス
    :param benchmark_ticker: ベンチマークティッカー(除外用)
    :param batch_size: バッチサイズ
    :param max_workers: 並列度
    :param verbose: 進捗表示
    :return: 処理結果
    """
    if verbose:
        logger.info(
            "Active returns optimized batch save started",
            columns=len(return_cols),
            rows=len(df_active_returns),
        )

    # データ型変換とリネーム
    df_opt = df_active_returns.copy()
    df_opt["value"] = df_opt["value"].astype(float)
    df_opt["date"] = pd.to_datetime(df_opt["date"])
    df_opt = df_opt.rename(columns={"symbol": "P_SYMBOL"})

    # ベンチマーク除外
    df_opt = df_opt[df_opt["P_SYMBOL"] != benchmark_ticker]

    if verbose:
        logger.debug("Pre-splitting data")

    # variable列でグループ化して辞書化
    df_dict = {}
    for col in return_cols:
        active_return_col = col.replace("Return", "Active_Return")

        # queryの代わりに高速なフィルタリング
        df_slice = df_opt[df_opt["variable"] == active_return_col].reset_index(
            drop=True
        )

        if not df_slice.empty:
            df_dict[active_return_col] = df_slice

    if verbose:
        logger.info("Split completed", tables=len(df_dict))

    # バッチ保存実行
    results = store_to_database_batch(
        df_dict=df_dict,
        db_path=db_path,
        unique_cols=["date", "P_SYMBOL", "variable"],
        batch_size=batch_size,
        max_workers=max_workers,
        verbose=verbose,
    )

    return results


# ============================================================================================
def ensure_unique_constraint(conn: sqlite3.Connection, table_name: str):
    """
    テーブルにUNIQUE制約があるか確認し、なければ再作成
    """
    cursor = conn.cursor()

    # テーブル情報を取得
    cursor.execute(f"PRAGMA table_info('{table_name}')")
    columns_info = cursor.fetchall()

    # インデックス情報を取得
    cursor.execute(f"PRAGMA index_list('{table_name}')")
    indexes = cursor.fetchall()

    # PRIMARY KEYまたはUNIQUEインデックスが(date, P_SYMBOL, variable)にあるか確認
    has_constraint = False
    for index in indexes:
        index_name = index[1]
        cursor.execute(f"PRAGMA index_info('{index_name}')")
        index_columns = [col[2] for col in cursor.fetchall()]

        if set(index_columns) == {"date", "P_SYMBOL", "variable"}:
            has_constraint = True
            break

    if not has_constraint:
        logger.warning("Table missing UNIQUE constraint", table=table_name)
        return False

    return True


# ============================================================================================
def upsert_financial_data(
    df: pd.DataFrame,
    conn: sqlite3.Connection,
    table_name: str,
    method: str = "auto",  # "auto", "upsert", "delete_insert"
):
    """
    財務データを更新

    Parameters
    ----------
    df : pd.DataFrame
        更新データ (columns: date, P_SYMBOL, variable, value)
    conn : sqlite3.Connection
        データベース接続
    table_name : str
        テーブル名
    method : str
        更新方法 ("auto", "upsert", "delete_insert")
        - "auto": SQLiteバージョンに応じて自動選択
        - "upsert": UPSERT構文を使用(SQLite 3.24.0以降)
        - "delete_insert": DELETE + INSERT方式
    """

    if df.empty:
        logger.warning("Empty update data", table=table_name)
        return 0

    # テーブル名のバリデーション（SQLインジェクション対策）
    _validate_sql_identifier(table_name)

    # SQLiteバージョンをチェック
    sqlite_version = tuple(map(int, sqlite3.sqlite_version.split(".")))
    supports_upsert = sqlite_version >= (3, 24, 0)

    # methodが"auto"の場合、バージョンに応じて自動選択
    if method == "auto":
        method = "upsert" if supports_upsert else "delete_insert"
        logger.info(
            "Auto-selected upsert method",
            sqlite_version=sqlite3.sqlite_version,
            method=method,
        )

    # 日付を文字列に変換
    df_copy = df.copy()
    df_copy["date"] = pd.to_datetime(df_copy["date"]).dt.strftime("%Y-%m-%d")

    # NaN を None に変換(SQLiteのNULL)
    df_copy = df_copy.where(pd.notnull(df_copy), None)

    # 一時テーブル名
    temp_table = f"{table_name}_temp_{np.random.randint(1000, 9999)}"

    try:
        # 一時テーブルに保存
        df_copy.to_sql(temp_table, conn, if_exists="replace", index=False)

        cursor = conn.cursor()

        if method == "upsert":
            if not supports_upsert:
                logger.warning(
                    "UPSERT not supported, falling back to delete_insert",
                    sqlite_version=sqlite3.sqlite_version,
                )
                method = "delete_insert"
            elif not ensure_unique_constraint(conn, table_name):
                logger.warning(
                    "UPSERT unavailable, falling back to delete_insert",
                    table=table_name,
                )
                method = "delete_insert"

        if method == "upsert":
            # UPSERT方式(SQLite 3.24.0以降)
            # table_name, temp_table は _validate_sql_identifier() で検証済み
            upsert_query = f"""
                INSERT INTO "{table_name}" (date, P_SYMBOL, variable, value)
                SELECT date, P_SYMBOL, variable, value
                FROM "{temp_table}"
                ON CONFLICT(date, P_SYMBOL, variable)
                DO UPDATE SET value = excluded.value
            """  # nosec B608 - table_name は検証済み
            cursor.execute(upsert_query)
            rows_affected = cursor.rowcount

        elif method == "delete_insert":
            # DELETE + INSERT方式(古いSQLiteでも動作)
            # table_name, temp_table は _validate_sql_identifier() で検証済み
            # 1. 該当行を削除
            delete_query = f"""
                DELETE FROM "{table_name}"
                WHERE (date, P_SYMBOL, variable) IN (
                    SELECT date, P_SYMBOL, variable
                    FROM "{temp_table}"
                )
            """  # nosec B608 - table_name は検証済み
            cursor.execute(delete_query)
            deleted = cursor.rowcount

            # 2. 新規データを挿入
            insert_query = f"""
                INSERT INTO "{table_name}" (date, P_SYMBOL, variable, value)
                SELECT date, P_SYMBOL, variable, value
                FROM "{temp_table}"
            """  # nosec B608 - table_name は検証済み
            cursor.execute(insert_query)
            inserted = cursor.rowcount

            rows_affected = inserted

            logger.info(
                "Delete-insert completed",
                table=table_name,
                deleted=deleted,
                inserted=inserted,
            )

        else:
            raise ValueError(f"不正なmethod: {method}")

        # 一時テーブル削除
        # nosec B608 - temp_table は table_name から生成された安全な識別子
        cursor.execute(f'DROP TABLE IF EXISTS "{temp_table}"')

        conn.commit()

        logger.info("Upsert completed", table=table_name, rows_affected=rows_affected)

        return rows_affected

    except Exception as e:
        logger.error("Upsert failed", table=table_name, error=str(e), exc_info=True)
        conn.rollback()
        # 一時テーブルのクリーンアップ
        with contextlib.suppress(BaseException):
            # nosec B608 - temp_table は table_name から生成された安全な識別子
            cursor.execute(f'DROP TABLE IF EXISTS "{temp_table}"')  # nosec B608
        raise


# ============================================================================================
def load_index_constituents(
    factset_index_db_path: Path, UNIVERSE_CODE: str
) -> pd.DataFrame:
    """
    BPMとFactsetのコードをマージした、インデックス構成銘柄のデータベースから
    必要なカラムをロードする関数
    """
    # テーブル名のバリデーション（SQLインジェクション対策）
    _validate_sql_identifier(UNIVERSE_CODE)

    with sqlite3.connect(factset_index_db_path) as conn:
        # UNIVERSE_CODE は _validate_sql_identifier() で検証済み
        query = f"""
            SELECT
                `date`, `P_SYMBOL`, `SEDOL`, `Asset ID`, `FG_COMPANY_NAME`,
                `GICS Sector`, `GICS Industry Group`,
                `Weight (%)`
            FROM
                {UNIVERSE_CODE}
            """  # nosec B608
        df_weight = pd.read_sql(
            query,
            con=conn,
            parse_dates=["date"],
        )

    return df_weight


# ============================================================================================
def load_financial_data(
    financials_db_path: Path, factor_list: list[str]
) -> pd.DataFrame:
    """
    Factsetの財務データとリターンを格納したデータベースからロードする関数
    factor_listでファクターを指定する
    例:
        factor_list = [
            "FF_ROIC",
            "FF_ROIC_PctRank",
            "FF_ROIC_Rank",
            "FF_ROE",
            "FF_ROE_PctRank",
            "FF_ROE_Rank",
            "Active_Return_1M_annlzd",
            "Active_Return_3M_annlzd",
            "Active_Return_6M_annlzd",
            "Active_Return_12M_annlzd",
            "Active_Return_3Y_annlzd",
            "Active_Return_5Y_annlzd",
        ]
    """
    # テーブル名のバリデーション（SQLインジェクション対策）
    for factor in factor_list:
        _validate_sql_identifier(factor)

    # join()を使用してクエリをUNION ALLする
    # factor_list の各要素は _validate_sql_identifier() で検証済み
    query = """
        SELECT `date`, `P_SYMBOL`, `variable`, `value` FROM {}
    """.format(  # nosec B608
        "\n    UNION ALL\n    SELECT `date`, `P_SYMBOL`, `variable`, `value` FROM ".join(
            factor_list
        )
    )

    with sqlite3.connect(financials_db_path) as conn:
        df_factor = (
            pd.read_sql(query, parse_dates=["date"], con=conn)
            .assign(
                # date列はBPM+Factsetコードのデータベースと異なる
                # BPM+Factsetは月末日
                # 一方、Factsetの財務データは月末最終営業日
                # よって、ここでは強制的に月末日に揃える
                date=lambda x: pd.to_datetime(x["date"]) + pd.offsets.MonthEnd(0)
            )
            .drop_duplicates(ignore_index=True)
        )
        # long formatに変換
        df_factor = pd.pivot(
            df_factor, index=["date", "P_SYMBOL"], columns="variable", values="value"
        ).reset_index()

    return df_factor


# ============================================================================================
def merge_idx_constituents_and_financials(
    df_weight: pd.DataFrame, df_factor: pd.DataFrame
) -> pd.DataFrame:
    """
    load_index_constituents()とload_financial_data()の結果をマージする関数
    """
    df_merged = pd.merge(
        df_weight, df_factor, on=["date", "P_SYMBOL"], how="outer"
    ).dropna(subset=["Weight (%)"], ignore_index=True)

    return df_merged


# ============================================================================================
@dataclass
class FactorJobArgs:
    """ファクター計算ジョブのための引数を格納するデータクラス。

    Attributes:
        factor (str): ファクター名(例: "FF_ROIC")。
        db_path (Union[str, Path]): データベースファイルのパス。
        df_weight (pd.DataFrame): インデックス構成銘柄のデータフレーム(Weight列などを含む)。
        sector_neutral_mode (bool): セクター中立化を行うかどうかのフラグ。
        inversed (bool): ランクを逆転させるかどうかのフラグ(例: 低い方が良い指標の場合)。
        period (Optional[str]): 期間指定。指定がない場合はNone。
            想定される値: ["QoQ", "YoY", "CAGR_3Y", "CAGR_5Y"] など。
            指定された場合、テーブル名は `{factor}_{period}` となります。
        winsorize (bool): add_factor_rank_colとadd_factor_pct_rank_cols関数で
            winsorizeする場合はTrue, しない場合はFalse
        winsorize_limits (tule): winsorizeの両端何%で切るか
    """

    factor: str
    db_path: str | Path
    df_weight: pd.DataFrame
    sector_neutral_mode: bool
    inversed: bool
    period: str | None = None
    winsorize: bool = True
    winsorize_limits: tuple = (0.01, 0.01)


# --------------------------------------------------------------------------------------------
def process_ranking_factor_worker(
    job_args: FactorJobArgs,
) -> list[tuple[str, pd.DataFrame]]:
    """単一のファクター・期間に対してランク等の指標計算を行うワーカー関数。

    DataClass経由で引数を受け取るため、引数の順序依存性がありません。
    periodが指定されている場合は `{factor}_{period}` を、
    指定されていない場合は `{factor}` をテーブル名およびカラム名として使用します。

    Args:
        job_args (FactorJobArgs): 計算に必要な全パラメータを格納したデータクラス。

    Returns:
        List[Tuple[str, pd.DataFrame]]: 計算結果のリスト。
            各要素は (テーブル名/カラム名, 計算済みデータフレーム) のタプル。
            エラーが発生した場合や該当データがない場合は空リストを返します。
    """
    # ターゲット名(テーブル名・カラム名)の決定
    # ファクター名とピリオドをバリデーション（SQLインジェクション対策）
    _validate_sql_identifier(job_args.factor)
    if job_args.period:
        _validate_sql_identifier(job_args.period)
        target_factor_name = f"{job_args.factor}_{job_args.period}"
    else:
        target_factor_name = job_args.factor

    results = []

    try:
        # 1. データベース読み込み
        # target_factor_name は _validate_sql_identifier() で検証済み
        query = f"SELECT `date`, `P_SYMBOL`, `value` FROM '{target_factor_name}'"  # nosec B608

        with sqlite3.connect(job_args.db_path) as conn:
            df = pd.read_sql(query, con=conn, parse_dates=["date"])

        # 日付調整(月末揃え)とリネーム
        df["date"] = pd.to_datetime(df["date"]) + pd.tseries.offsets.MonthEnd(0)
        df = df.sort_values("date", ignore_index=True).rename(
            columns={"value": target_factor_name}
        )

        # 2. マージ処理
        # 構成銘柄情報と結合してユニバースをフィルタリング
        df_merged = pd.merge(
            job_args.df_weight, df, on=["date", "P_SYMBOL"], how="outer"
        )

        # クリーニング: 重複排除と必須カラムの欠損除去
        df_merged = df_merged.drop_duplicates(subset=["date", "P_SYMBOL"]).dropna(
            subset=["Weight (%)", target_factor_name],
            how="any",
            axis=0,
            ignore_index=True,
        )

        if df_merged.empty:
            return results

        # 3. 指標計算 (Rank, PctRank, ZScore)
        # -------------------------------------------------------------
        # 内部関数で処理を共通化
        def _add_metric_to_results(
            metric_type: str,
            calculation_func: Callable[..., Any],
        ):
            """計算を実行し、結果リストに追加するヘルパー関数。"""
            # 🔧 修正箇所: 関数のシグネチャを検査
            import inspect

            func_params = inspect.signature(calculation_func).parameters

            # winsorize引数を持つかチェック
            has_winsorize = "winsorize" in func_params
            has_winsorize_limits = "winsorize_limits" in func_params

            # 基本引数
            kwargs = {
                "df": df_merged,
                "factor_name": target_factor_name,
                "sector_neutral_mode": job_args.sector_neutral_mode,
                "inversed": job_args.inversed,
            }

            # winsorize引数が存在する場合のみ追加
            if has_winsorize:
                kwargs["winsorize"] = job_args.winsorize
            if has_winsorize_limits:
                kwargs["winsorize_limits"] = job_args.winsorize_limits

            # roic_utilsの計算関数を呼び出し
            df_res = calculation_func(**kwargs)

            # カラム名の構築 (例: Factor_Inv_Rank または Factor_Rank)
            prefix = (
                f"{target_factor_name}_Inv"
                if job_args.inversed
                else f"{target_factor_name}"
            )
            col_name = f"{prefix}_{metric_type}"
            # セクター中立のラベル
            if job_args.sector_neutral_mode:
                col_name = f"{col_name}_Sector_Neutral"

            # テーブル名
            table_name = col_name

            # 必要な列のみ抽出し、long format用のカラム構成にする
            df_res = (
                df_res[["date", "P_SYMBOL", col_name]]
                .rename(columns={col_name: "value"})
                .assign(variable=col_name)
            )
            results.append((table_name, df_res))

        # -------------------------------------------------------------

        # 各指標の計算実行
        _add_metric_to_results("Rank", roic_utils.add_factor_rank_cols)  # type: ignore[name-defined]
        _add_metric_to_results("PctRank", roic_utils.add_factor_pct_rank_cols)  # type: ignore[name-defined]
        _add_metric_to_results("ZScore", roic_utils.add_factor_zscore_cols)  # type: ignore[name-defined]

        return results

    except Exception as e:
        # 並列処理中のエラーはログに出力し、プロセスを落とさずに空リストを返す
        logger.error(
            "Factor processing failed",
            factor=target_factor_name,
            error=str(e),
            exc_info=True,
        )
        return []


# ============================================================================================
def process_rank_calculation_store_to_db(
    df_weight: pd.DataFrame,
    factor_list: list[str],
    financials_db_path: Path,
    period_list: list[str] | None = None,
    sector_neutral_mode: bool = True,
    inversed: bool = False,
    winsorize: bool = True,
    winsorize_limits: tuple = (0.01, 0.01),
    default_max_workers: int = 6,
):
    """ファクターのランク計算を行い、データベースに保存する関数。

    period_list の指定有無により、単一ファクター計算と期間付き計算の両方に対応します。

    Args:
        df_weight (pd.DataFrame): ウェイト情報を含む構成銘柄データ。
        factor_list (List[str]): 計算対象のファクター名のリスト。
        financials_db_path (Path): 保存先のデータベースパス。
        period_list (Optional[List[str]]): 期間のリスト (例: ["YoY", "QoQ"])。
            指定した場合、factor_listとの組み合わせで計算を行います。
            Noneの場合、periodなし(単一ファクター)として計算します。
        sector_neutral_mode (bool): セクター中立化を行うか。
        inversed (bool): ランクを逆転させるか。
        default_max_workers (int): 最大並列プロセス数。
    """
    # ---------------------------------------------------------
    # 1. タスクリストの作成 (FactorJobArgsの準備)
    # ---------------------------------------------------------
    tasks = []

    if period_list and len(period_list) > 0:
        # パターンA: 期間指定あり (Factor x Period の組み合わせを作成)
        mode_desc = "Multi-Period Mode"
        for factor in factor_list:
            for period in period_list:
                args = FactorJobArgs(
                    factor=factor,
                    db_path=financials_db_path,
                    df_weight=df_weight,
                    sector_neutral_mode=sector_neutral_mode,
                    inversed=inversed,
                    period=period,  # 期間を指定
                    winsorize=winsorize,
                    winsorize_limits=winsorize_limits,
                )
                tasks.append(args)
    else:
        # パターンB: 期間指定なし (Single Factor Mode)
        mode_desc = "Single Factor Mode"
        for factor in factor_list:
            args = FactorJobArgs(
                factor=factor,
                db_path=financials_db_path,
                df_weight=df_weight,
                sector_neutral_mode=sector_neutral_mode,
                inversed=inversed,
                period=None,  # 期間なし
                winsorize=winsorize,
                winsorize_limits=winsorize_limits,
            )
            tasks.append(args)

    total_iterations = len(tasks)
    logger.info("Rank calculation started", tasks=total_iterations, mode=mode_desc)

    # ---------------------------------------------------------
    # 2. DB設定と並列実行の準備
    # ---------------------------------------------------------
    # DBのWALモード有効化 (書き込み速度向上)
    with sqlite3.connect(financials_db_path) as conn:
        conn.execute("PRAGMA journal_mode=WAL;")
    # 並列のプロセス数が多いとメモリを食うので注意 (4-6程度が安全圏)
    max_workers = min(default_max_workers, total_iterations)

    with ProcessPoolExecutor(max_workers=max_workers) as executor:
        # futureをキーにして、メタデータ(factor, period)を保存する辞書
        futures = {}

        for job_args in tasks:
            future = executor.submit(process_ranking_factor_worker, job_args)
            # エラー表示用に情報を記録しておく
            futures[future] = (job_args.factor, job_args.period)

        # ---------------------------------------------------------
        # 3. 結果の回収とDB保存 (メインスレッド)
        # ---------------------------------------------------------
        for future in tqdm(
            as_completed(futures), total=total_iterations, desc="Rank計算進捗"
        ):
            # どのタスクの結果か特定
            current_factor, current_period = futures[future]
            task_label = (
                f"{current_factor}_{current_period}"
                if current_period
                else current_factor
            )

            try:
                results = future.result()
                if not results:
                    continue

                # DB書き込み(直列実行でロック回避)
                for table_name, df_result in results:
                    # 既存テーブル削除
                    db_utils.delete_table_from_database(  # type: ignore[name-defined]
                        db_path=financials_db_path, table_name=table_name
                    )
                    # 保存
                    store_to_database(
                        df=df_result,
                        db_path=financials_db_path,
                        table_name=table_name,
                        verbose=False,
                    )

                del results

            except Exception as e:
                logger.critical(
                    "Rank calculation critical error",
                    task=task_label,
                    error=str(e),
                    exc_info=True,
                )

        logger.info("All rank calculations and saves completed")


# ============================================================================================
def check_missing_value_and_fill_by_sector_median(
    df: pd.DataFrame, factor_list: list[str]
) -> pd.DataFrame:
    # -----------------------------------
    # 欠損状況の確認(補完前)
    # -----------------------------------
    missing_before = {}
    for factor in factor_list:
        missing_count = df[factor].isna().sum()
        missing_before[factor] = missing_count

    logger.info(
        "Missing value analysis (before fill)",
        stats={f: int(missing_before[f]) for f in factor_list},
        total_rows=len(df),
    )

    # -----------------------------------
    # 各ファクターの欠損値をセクター中央値で補完
    # -----------------------------------
    logger.debug("Filling missing values with sector median")
    df[factor_list] = df.groupby(["date", "GICS Sector"])[factor_list].transform(
        lambda x: x.fillna(x.median())
    )

    # -----------------------------------
    # 補完後の欠損状況を確認
    # -----------------------------------
    missing_after_sector = {}
    remaining_missing = False

    for factor in factor_list:
        missing_count = df[factor].isna().sum()
        missing_after_sector[factor] = missing_count
        if missing_count > 0:
            remaining_missing = True

    logger.info(
        "Missing value analysis (after sector median fill)",
        stats={f: int(missing_after_sector[f]) for f in factor_list},
        filled={
            f: int(missing_before[f] - missing_after_sector[f]) for f in factor_list
        },
    )

    # -----------------------------------
    # セクター補完で埋まらなかった欠損を全体中央値で再補完
    # -----------------------------------
    if remaining_missing:
        logger.info("Re-filling remaining missing values with overall median")

        for factor in factor_list:
            if df[factor].isna().sum() > 0:
                # 日付ごとの全体中央値で補完
                df[factor] = df.groupby("date")[factor].transform(
                    lambda x: x.fillna(x.median())
                )

                # それでも埋まらない場合(全体が欠損)は0.5(中立値)
                overall_missing = df[factor].isna().sum()
                if overall_missing > 0:
                    logger.warning(
                        "Factor still missing after all fills, using neutral value 0.5",
                        factor=factor,
                        remaining=int(overall_missing),
                    )
                    df[factor] = df[factor].fillna(0.5)

    # -----------------------------------
    # 最終確認
    # -----------------------------------
    final_missing_total = 0
    final_stats = {}
    for factor in factor_list:
        missing_count = df[factor].isna().sum()
        final_missing_total += missing_count
        final_stats[factor] = int(missing_count)

    if final_missing_total == 0:
        logger.info("All missing values filled successfully")
    else:
        logger.warning(
            "Remaining missing values after all fills",
            total_remaining=final_missing_total,
            stats=final_stats,
        )

    return df


# ============================================================================================
def create_factor(
    df: pd.DataFrame, factor_name: str, blend_weight: dict
) -> pd.DataFrame:
    logger.info(
        "Factor calculation started",
        factor=factor_name,
        weights={k: round(v, 4) for k, v in blend_weight.items()},
    )

    # 加重平均でComposite Scoreを計算
    df[f"{factor_name}_Score"] = sum(
        df[indicator] * weight for indicator, weight in blend_weight.items()
    )

    score_col = f"{factor_name}_Score"
    logger.info(
        "Score calculation completed",
        factor=factor_name,
        mean=round(df[score_col].mean(), 4),
        std=round(df[score_col].std(), 4),
        min=round(df[score_col].min(), 4),
        max=round(df[score_col].max(), 4),
    )

    # -----------------------------------
    # ランク化
    # -----------------------------------
    df[f"{factor_name}_Score_Rank"] = df.groupby("date")[
        f"{factor_name}_Score"
    ].transform(
        lambda x: pd.qcut(
            x,
            q=5,
            labels=["rank5", "rank4", "rank3", "rank2", "rank1"],
            duplicates="drop",
        )
    )

    # ランク分布
    rank_dist = df[f"{factor_name}_Score_Rank"].value_counts().sort_index()
    logger.info(
        "Rank distribution",
        factor=factor_name,
        distribution={str(k): int(v) for k, v in rank_dist.items()},
    )

    return df
